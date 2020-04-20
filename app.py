#!/usr/bin/env python3

import os
import time
import openpyxl
from flask import Flask, request, render_template

app = Flask(__name__, template_folder = 'templates', static_folder = 'static')
app.config.update(
	DEBUG = True,
	SERVER_NAME = '127.0.0.1:2019',
	UPLOAD_FOLDER = 'storage'
)


@app.route('/', methods = ['GET'])
def index():
	if not os.path.exists(app.config['UPLOAD_FOLDER']):
		os.makedirs(app.config['UPLOAD_FOLDER'])
	return render_template('index.j2')

@app.route('/upload', methods = ['POST'])
def upload():
	file = request.files['file']
	allowed_extensions = ['xlsx', 'xls', 'xlsm']
	if file.filename.rsplit('.')[-1] not in allowed_extensions:
		return '', 400
	full_filename = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
	file.save(full_filename)
	return '', 200

@app.route('/uploaded_files', methods = ['GET'])
def uploaded_files():
	files = []
	for filename in os.listdir(app.config['UPLOAD_FOLDER']):
		full_filename = os.path.join(app.config['UPLOAD_FOLDER'], filename)
		files.append({
			'filename': filename,
			'creation_time': time.ctime(os.stat(full_filename).st_atime)
		})
	return render_template('files_list.j2', files=files)

@app.route('/view_file_content', methods = ['GET'])
def view_file_content():
	filename = request.args.get('filename')
	wb = openpyxl.load_workbook(os.path.join(app.config['UPLOAD_FOLDER'], filename))
	sheet = wb.active
	# read all data from xlsx active book to list of lists
	table = []
	for row in sheet.rows:
		tmp_row = []
		for cell in row:
			tmp_row.append(cell.value if cell.value!=None else '-' )
		table.append(tmp_row)
	# collect string column names; transform 1,2,3,... to A,B,C,...
	cols_ids = []
	for i in range(1, sheet.max_column + 1):
		cols_ids.append(openpyxl.utils.get_column_letter(i))
	return render_template('xlsx_content.j2', table=table, cols_ids=cols_ids)


if __name__ == '__main__':
	app.run()
